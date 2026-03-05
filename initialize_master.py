"""
BOACTAR Master File Initialization
===================================
One-time setup script to initialize the master data file from historical data.

Run this script once before using the main pipeline:
    python initialize_master.py
"""

import os
import shutil
import logging

import config
from logger_setup import setup_logging

logger = logging.getLogger(__name__)


def initialize_master_from_historical():
    """
    Initialize master data file from historical data file.

    Copies the historical data file to the master data location.
    """
    print('BOACTAR Master File Initialization')
    print('=' * 50)

    # Check if historical file exists
    if not os.path.exists(config.HISTORICAL_DATA_FILE):
        print(f'ERROR: Historical data file not found:')
        print(f'  {config.HISTORICAL_DATA_FILE}')
        return False

    # Check if master already exists
    if os.path.exists(config.MASTER_DATA_FILE):
        print(f'Master file already exists:')
        print(f'  {config.MASTER_DATA_FILE}')

        response = input('Overwrite? (y/N): ').strip().lower()
        if response != 'y':
            print('Aborted.')
            return False

    # Create master directory if needed
    os.makedirs(config.MASTER_DIR, exist_ok=True)

    # Copy historical file to master location
    print(f'\nCopying historical data to master file...')
    print(f'  From: {config.HISTORICAL_DATA_FILE}')
    print(f'  To:   {config.MASTER_DATA_FILE}')

    shutil.copy2(config.HISTORICAL_DATA_FILE, config.MASTER_DATA_FILE)

    # Verify
    if os.path.exists(config.MASTER_DATA_FILE):
        size = os.path.getsize(config.MASTER_DATA_FILE)
        print(f'\nMaster file created successfully!')
        print(f'  Size: {size:,} bytes')

        # Read and display info
        import openpyxl
        wb = openpyxl.load_workbook(config.MASTER_DATA_FILE)
        ws = wb.active
        print(f'  Rows: {ws.max_row}')
        print(f'  Columns: {ws.max_column}')

        # Get date range
        dates = []
        for row in range(3, ws.max_row + 1):
            date_val = ws.cell(row=row, column=1).value
            if date_val:
                dates.append(date_val)

        if dates:
            print(f'  Date range: {min(dates)} to {max(dates)}')
            print(f'  Total data rows: {len(dates)}')

        wb.close()

        return True
    else:
        print('ERROR: Failed to create master file')
        return False


def main():
    """Main entry point."""
    setup_logging()

    success = initialize_master_from_historical()

    if success:
        print('\n' + '=' * 50)
        print('Initialization complete!')
        print('You can now run: python orchestrator.py')
        print('=' * 50)
        return 0
    else:
        return 1


if __name__ == '__main__':
    import sys
    sys.exit(main())
