"""
BOACTAR Data Loader
===================
Loads and cleans input Excel files from the input folder.

This module handles:
- Finding input files (.xlsm, .xlsx) in the input directory
- Reading the Summary tab from input files
- Extracting the data date from the file
- Parsing all data rows with dynamic header detection
- Cleaning data (converting percentages, handling NA values)
"""

import os
import logging
from datetime import datetime, timedelta
from glob import glob

import openpyxl

import config

logger = logging.getLogger(__name__)


class BOACTARDataLoader:
    """Loads and cleans BOA CTA Report Excel files."""

    def __init__(self):
        """Initialize the data loader."""
        self.input_dir = config.INPUT_DIR

    def find_input_files(self):
        """
        Find all input files in the input directory.

        Returns:
            list: List of paths to input files (.xlsm, .xlsx).
        """
        patterns = ['*.xlsm', '*.xlsx']
        files = []

        for pattern in patterns:
            search_path = os.path.join(self.input_dir, pattern)
            found = glob(search_path)
            files.extend(found)

        # Sort by modification time (newest first)
        files.sort(key=os.path.getmtime, reverse=True)

        logger.info(f'Found {len(files)} input file(s) in {self.input_dir}')
        for f in files:
            logger.debug(f'  - {os.path.basename(f)}')

        return files

    def _excel_date_to_datetime(self, excel_date):
        """
        Convert Excel serial date number to datetime.

        Args:
            excel_date: Excel serial date number or datetime object.

        Returns:
            datetime: Converted datetime object.
        """
        if isinstance(excel_date, datetime):
            return excel_date

        if isinstance(excel_date, (int, float)):
            # Excel date serial number
            return config.EXCEL_DATE_EPOCH + timedelta(days=int(excel_date))

        # Try parsing as string
        if isinstance(excel_date, str):
            for fmt in config.INPUT_DATE_FORMATS:
                try:
                    return datetime.strptime(excel_date, fmt)
                except ValueError:
                    continue

        raise ValueError(f'Unable to parse date: {excel_date}')

    def _clean_value(self, value, field_key=None):
        """
        Clean a single cell value.

        Handles:
        - NA values (dashes, empty cells)
        - Percentage conversion (multiply by 100 for percentage fields)
        - Preserving full precision for financial data

        Args:
            value: Raw cell value.
            field_key: Optional field key to determine if percentage conversion needed.

        Returns:
            Cleaned value (float, str 'NA', or original value).
        """
        # Check for NA values
        if value in config.NA_INPUT_VALUES:
            return config.NA_VALUE

        # Handle string dashes
        if isinstance(value, str):
            stripped = value.strip()
            if stripped in ['-', '--', '']:
                return config.NA_VALUE

        # Handle numeric values
        if isinstance(value, (int, float)):
            # Check if this is a percentage field that needs conversion
            if field_key and config.is_percentage_field(field_key):
                # Multiply by 100 to convert from decimal to percentage
                # Only if value looks like a decimal (between -1 and 1 range typically)
                # But some values can be > 1, so we always multiply
                return value * 100
            return value

        return value

    def _detect_asset_class(self, ws, row):
        """
        Detect the asset class for a row.

        The asset class is in column B but only appears once per group.
        We need to track and carry forward the asset class.

        Args:
            ws: Worksheet object.
            row: Current row number.

        Returns:
            str: Asset class name or None.
        """
        col = config.INPUT_COLUMNS['ASSET_CLASS']
        value = ws.cell(row=row, column=col).value
        if value and str(value).strip():
            return str(value).strip()
        return None

    def _get_underlying(self, ws, row):
        """
        Get the underlying name from a row.

        Args:
            ws: Worksheet object.
            row: Current row number.

        Returns:
            str: Underlying name or None.
        """
        col = config.INPUT_COLUMNS['UNDERLYING']
        value = ws.cell(row=row, column=col).value
        if value and str(value).strip():
            return str(value).strip()
        return None

    def _extract_row_data(self, ws, row, current_asset_class):
        """
        Extract all data from a single row.

        Args:
            ws: Worksheet object.
            row: Row number to extract.
            current_asset_class: Current asset class context.

        Returns:
            dict: Dictionary with underlying name and all field values, or None if empty row.
        """
        underlying = self._get_underlying(ws, row)
        if not underlying:
            return None

        # Check for new asset class
        asset_class = self._detect_asset_class(ws, row)
        if asset_class:
            current_asset_class = asset_class

        # Extract all field values
        row_data = {
            'underlying': underlying,
            'asset_class': current_asset_class,
        }

        # Extract each field
        for field_key, col_num in config.INPUT_COLUMNS.items():
            if field_key in ['ASSET_CLASS', 'UNDERLYING']:
                continue

            raw_value = ws.cell(row=row, column=col_num).value
            cleaned_value = self._clean_value(raw_value, field_key)
            row_data[field_key] = cleaned_value

        return row_data, current_asset_class

    def load_file(self, file_path):
        """
        Load and parse a single input file.

        Args:
            file_path: Path to the input file.

        Returns:
            dict: Dictionary containing:
                - 'date': datetime object for the data date
                - 'date_str': Formatted date string
                - 'file_name': Input file name
                - 'data': List of row data dictionaries
        """
        logger.info(f'Loading file: {os.path.basename(file_path)}')

        try:
            # Load workbook with data_only=True to get calculated values
            wb = openpyxl.load_workbook(file_path, data_only=True)

            # Check for Summary sheet
            if config.INPUT_SHEET_NAME not in wb.sheetnames:
                raise ValueError(f"Sheet '{config.INPUT_SHEET_NAME}' not found in {file_path}")

            ws = wb[config.INPUT_SHEET_NAME]

            # Extract date from the date cell
            date_row, date_col = config.INPUT_DATE_CELL
            date_value = ws.cell(row=date_row, column=date_col).value
            data_date = self._excel_date_to_datetime(date_value)
            date_str = data_date.strftime(config.DATE_FORMAT_OUTPUT)

            logger.info(f'Data date: {date_str}')

            # Extract all data rows
            data_rows = []
            current_asset_class = None

            row = config.INPUT_DATA_START_ROW
            max_row = ws.max_row

            while row <= max_row:
                result = self._extract_row_data(ws, row, current_asset_class)

                if result is not None:
                    row_data, current_asset_class = result
                    data_rows.append(row_data)
                    logger.debug(f"Row {row}: {row_data['asset_class']} - {row_data['underlying']}")

                row += 1

            logger.info(f'Extracted {len(data_rows)} data rows')

            wb.close()

            return {
                'date': data_date,
                'date_str': date_str,
                'file_name': os.path.basename(file_path),
                'file_path': file_path,
                'data': data_rows,
            }

        except Exception as e:
            logger.error(f'Error loading file {file_path}: {e}')
            raise

    def load_latest_file(self):
        """
        Load the most recent input file.

        Returns:
            dict: Loaded data dictionary or None if no files found.
        """
        files = self.find_input_files()

        if not files:
            logger.warning('No input files found')
            return None

        # Load the most recent file
        return self.load_file(files[0])

    def load_all_files(self, sort_by_date=True):
        """
        Load all input files in the input directory.

        Args:
            sort_by_date: If True, sort files by their data date (oldest first).

        Returns:
            list: List of loaded data dictionaries, sorted by date (oldest first).
        """
        files = self.find_input_files()
        results = []

        for file_path in files:
            try:
                data = self.load_file(file_path)
                results.append(data)
            except Exception as e:
                logger.error(f'Failed to load {file_path}: {e}')
                if not config.CONTINUE_ON_ERROR:
                    raise

        # Sort by data date (oldest first) so we process chronologically
        if sort_by_date and results:
            results.sort(key=lambda x: x['date'])
            logger.info(f'Sorted {len(results)} files by date (oldest to newest)')
            for data in results:
                logger.info(f"  - {data['date_str']}: {data['file_name']}")

        return results


def load_data():
    """
    Convenience function to load the latest input file.

    Returns:
        dict: Loaded data dictionary.
    """
    loader = BOACTARDataLoader()
    return loader.load_latest_file()


if __name__ == '__main__':
    # Test the data loader
    from logger_setup import setup_logging

    setup_logging()

    loader = BOACTARDataLoader()

    # Find files
    files = loader.find_input_files()
    print(f'\nFound {len(files)} input files')

    if files:
        # Load the first file
        data = loader.load_file(files[0])
        print(f"\nLoaded data from: {data['file_name']}")
        print(f"Date: {data['date_str']}")
        print(f"Total rows: {len(data['data'])}")

        # Print first few rows
        print('\nFirst 5 rows:')
        for i, row in enumerate(data['data'][:5]):
            print(f"  {i+1}. {row['asset_class']} - {row['underlying']}")
            print(f"      Current Price: {row.get('CURRENT_PRICE')}")
            print(f"      Short Term Strength: {row.get('SHORT_TERM_STRENGTH')}")
