"""
BOACTAR Data Parser
===================
Parses cleaned input data and merges with master data file.

This module handles:
- Mapping input data to output column codes
- Loading existing master data
- Checking if new data is more recent
- Merging new data rows with historical data
"""

import os
import logging
from datetime import datetime

import pandas as pd
import openpyxl

import config

logger = logging.getLogger(__name__)


class BOACTARParser:
    """Parses and merges BOA CTA Report data."""

    def __init__(self):
        """Initialize the parser."""
        self.column_order = None
        self._load_column_order()

    def _load_column_order(self):
        """
        Load the column order from the master file.

        This ensures output columns match the exact order in the master file.
        """
        if os.path.exists(config.MASTER_DATA_FILE):
            try:
                wb = openpyxl.load_workbook(config.MASTER_DATA_FILE)
                ws = wb.active

                # Read column codes from row 1 (skip first column which is dates)
                codes = []
                descs = []
                for col in range(2, ws.max_column + 1):
                    code = ws.cell(row=1, column=col).value
                    desc = ws.cell(row=2, column=col).value
                    if code:
                        codes.append(code)
                        descs.append(desc)

                self.column_order = list(zip(codes, descs))
                logger.info(f'Loaded {len(self.column_order)} columns from master file')
                wb.close()

            except Exception as e:
                logger.warning(f'Could not load column order from master: {e}')
                self.column_order = None

        # Fallback: load from historical file
        if self.column_order is None and os.path.exists(config.HISTORICAL_DATA_FILE):
            try:
                wb = openpyxl.load_workbook(config.HISTORICAL_DATA_FILE)
                ws = wb.active

                codes = []
                descs = []
                for col in range(2, ws.max_column + 1):
                    code = ws.cell(row=1, column=col).value
                    desc = ws.cell(row=2, column=col).value
                    if code:
                        codes.append(code)
                        descs.append(desc)

                self.column_order = list(zip(codes, descs))
                logger.info(f'Loaded {len(self.column_order)} columns from historical file')
                wb.close()

            except Exception as e:
                logger.error(f'Could not load column order from historical file: {e}')
                raise

    def _map_row_to_columns(self, row_data):
        """
        Map a single row of input data to output column values.

        Args:
            row_data: Dictionary with underlying name and field values.

        Returns:
            dict: Mapping of column code to value.
        """
        underlying = row_data['underlying']
        canonical_name = config.normalize_underlying_name(underlying)

        if canonical_name not in config.ASSET_MAPPING:
            logger.warning(f'Unknown underlying: {underlying}')
            return {}

        category, spot_code, trend_code, display_name = config.ASSET_MAPPING[canonical_name]
        result = {}

        # Map Current Price to SPOT column
        spot_col_code = f'BOACTAR.{category}.{spot_code}.SPOT.B'
        current_price = row_data.get('CURRENT_PRICE')
        if current_price is not None:
            result[spot_col_code] = current_price

        # Map trend fields
        for field_key, trend_field_code in config.INPUT_TO_TREND_FIELD.items():
            value = row_data.get(field_key)
            if value is not None:
                col_code = f'BOACTAR.{category}.{trend_code}.{trend_field_code}.B'
                result[col_code] = value

        return result

    def parse_loaded_data(self, loaded_data):
        """
        Parse loaded data into a row dictionary matching output columns.

        Args:
            loaded_data: Dictionary from data_loader with date and data rows.

        Returns:
            dict: Single row dictionary with date and all column values.
        """
        date_str = loaded_data['date_str']
        data_rows = loaded_data['data']

        # Initialize result with all columns set to NA
        result = {'date': date_str}

        if self.column_order:
            for code, desc in self.column_order:
                result[code] = config.NA_VALUE

        # Process each data row
        mapped_count = 0
        for row_data in data_rows:
            column_values = self._map_row_to_columns(row_data)
            result.update(column_values)
            if column_values:
                mapped_count += 1

        logger.info(f'Mapped {mapped_count} assets to output columns')

        return result

    def load_master_data(self):
        """
        Load existing master data file.

        Returns:
            pd.DataFrame: Master data as DataFrame, or empty DataFrame if not exists.
        """
        if not os.path.exists(config.MASTER_DATA_FILE):
            logger.info('Master data file not found, will create new')
            return pd.DataFrame()

        try:
            logger.info(f'Loading master data from {config.MASTER_DATA_FILE}')

            wb = openpyxl.load_workbook(config.MASTER_DATA_FILE)
            ws = wb.active

            # Read column codes from row 1 and determine actual column count
            columns = ['date']
            num_data_cols = 0
            for col in range(2, ws.max_column + 1):
                code = ws.cell(row=1, column=col).value
                if code:
                    columns.append(code)
                    num_data_cols = col  # Track last column with a code

            # Total columns to read (date column + data columns)
            total_cols = num_data_cols

            # Read data starting from row 3, only up to the columns we have headers for
            data = []
            for row in range(3, ws.max_row + 1):
                row_data = []
                for col in range(1, total_cols + 1):
                    value = ws.cell(row=row, column=col).value
                    row_data.append(value)
                if row_data[0]:  # Only add rows with a date
                    data.append(row_data)

            wb.close()

            # Create DataFrame
            df = pd.DataFrame(data, columns=columns)

            # Convert date column
            df['date'] = pd.to_datetime(df['date'], errors='coerce')
            df = df.dropna(subset=['date'])
            df = df.sort_values('date')

            logger.info(f'Loaded {len(df)} rows from master data')
            logger.info(f'Date range: {df["date"].min()} to {df["date"].max()}')

            return df

        except Exception as e:
            logger.error(f'Error loading master data: {e}')
            raise

    def get_latest_date_in_master(self, master_df):
        """
        Get the most recent date in the master data.

        Args:
            master_df: Master data DataFrame.

        Returns:
            datetime: Latest date or None if empty.
        """
        if master_df.empty:
            return None

        return master_df['date'].max()

    def is_new_data(self, data_date, latest_master_date):
        """
        Check if the new data is more recent than master.

        Args:
            data_date: Date of the new data.
            latest_master_date: Latest date in master.

        Returns:
            bool: True if new data should be added.
        """
        if latest_master_date is None:
            return True

        # Convert to datetime if needed
        if isinstance(data_date, str):
            data_date = datetime.strptime(data_date, config.DATE_FORMAT_OUTPUT)

        if isinstance(latest_master_date, str):
            latest_master_date = datetime.strptime(latest_master_date, config.DATE_FORMAT_OUTPUT)

        # Normalize to date only (remove time component)
        data_date = pd.Timestamp(data_date).normalize()
        latest_master_date = pd.Timestamp(latest_master_date).normalize()

        return data_date > latest_master_date

    def merge_data(self, master_df, new_row_data):
        """
        Merge new data row with master DataFrame.

        Args:
            master_df: Existing master DataFrame.
            new_row_data: Dictionary with date and column values.

        Returns:
            pd.DataFrame: Combined DataFrame.
        """
        # Convert new row to DataFrame
        date_str = new_row_data.pop('date')
        new_date = pd.to_datetime(date_str)

        # Build new row with all columns from master
        if not master_df.empty:
            new_row = {col: config.NA_VALUE for col in master_df.columns}
        else:
            new_row = {'date': new_date}
            for code, desc in self.column_order:
                new_row[code] = config.NA_VALUE

        new_row['date'] = new_date
        new_row.update(new_row_data)

        new_df = pd.DataFrame([new_row])

        # Combine with master
        if master_df.empty:
            combined = new_df
        else:
            combined = pd.concat([master_df, new_df], ignore_index=True)

        # Remove duplicates by date (keep last)
        combined = combined.drop_duplicates(subset=['date'], keep='last')

        # Sort by date
        combined = combined.sort_values('date')

        logger.info(f'Combined data has {len(combined)} rows')

        return combined

    def parse_and_merge(self, loaded_data, existing_df=None):
        """
        Main method to parse loaded data and merge with master.

        Args:
            loaded_data: Dictionary from data_loader.
            existing_df: Optional existing DataFrame to merge into (for batch processing).
                         If None, loads from master file.

        Returns:
            pd.DataFrame: Combined data ready for output.
        """
        # Parse the loaded data
        new_row_data = self.parse_loaded_data(loaded_data)
        data_date = loaded_data['date']

        # Use existing DataFrame if provided, otherwise load master data
        if existing_df is not None:
            master_df = existing_df
            logger.info(f'Using existing DataFrame with {len(master_df)} rows')
        else:
            master_df = self.load_master_data()

        # Check if data is new
        latest_master_date = self.get_latest_date_in_master(master_df)

        if latest_master_date:
            logger.info(f'Latest date in master: {latest_master_date}')
            logger.info(f'New data date: {data_date}')

        if not self.is_new_data(data_date, latest_master_date) and not config.REBUILD_MASTER:
            logger.info('Data is not newer than master, skipping merge')
            return master_df

        # Merge new data
        logger.info('Merging new data with master')
        combined_df = self.merge_data(master_df, new_row_data)

        return combined_df

    def get_column_order(self):
        """
        Get the column order for output files.

        Returns:
            list: List of (code, description) tuples.
        """
        return self.column_order


def parse_and_merge(loaded_data, existing_df=None):
    """
    Convenience function to parse and merge data.

    Args:
        loaded_data: Dictionary from data_loader.
        existing_df: Optional existing DataFrame to merge into.

    Returns:
        pd.DataFrame: Combined data.
    """
    parser = BOACTARParser()
    return parser.parse_and_merge(loaded_data, existing_df=existing_df)


if __name__ == '__main__':
    # Test the parser
    from logger_setup import setup_logging
    from data_loader import load_data

    setup_logging()

    # Load data
    loaded_data = load_data()

    if loaded_data:
        print(f"\nLoaded data from: {loaded_data['file_name']}")
        print(f"Date: {loaded_data['date_str']}")

        # Parse and merge
        parser = BOACTARParser()
        combined_df = parser.parse_and_merge(loaded_data)

        print(f'\nCombined DataFrame shape: {combined_df.shape}')
        print(f'Date range: {combined_df["date"].min()} to {combined_df["date"].max()}')

        # Show sample columns
        print('\nSample columns:')
        for col in list(combined_df.columns)[:10]:
            print(f'  {col}')
    else:
        print('No data loaded')
