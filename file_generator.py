"""
BOACTAR File Generator
======================
Generates output DATA, META, and ZIP files.

This module handles:
- Creating DATA Excel files with proper header structure
- Creating META Excel files with time series metadata
- Creating ZIP archives containing DATA and META files
- Updating the master data file
- Managing timestamped and latest output folders
"""

import os
import logging
import shutil
import zipfile
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

import config

logger = logging.getLogger(__name__)


class BOACTARFileGenerator:
    """Generates output files for BOACTAR runbook."""

    def __init__(self, column_order=None):
        """
        Initialize the file generator.

        Args:
            column_order: List of (code, description) tuples for columns.
        """
        self.column_order = column_order
        self.timestamp = config.get_timestamp()

    def _create_output_dirs(self):
        """Create output directory structure."""
        # Timestamped directory
        ts_dir = os.path.join(config.OUTPUT_DIR, self.timestamp)
        os.makedirs(ts_dir, exist_ok=True)

        # Latest directory
        latest_dir = os.path.join(config.OUTPUT_DIR, config.LATEST_FOLDER)
        os.makedirs(latest_dir, exist_ok=True)

        # Master directory
        os.makedirs(config.MASTER_DIR, exist_ok=True)

        return ts_dir, latest_dir

    def create_data_file(self, df, output_path):
        """
        Create DATA Excel file with proper structure.

        Structure:
        - Row 1: Column codes (empty first cell, then codes)
        - Row 2: Column descriptions (empty first cell, then descriptions)
        - Row 3+: Date and data values

        Args:
            df: DataFrame with data (date column + all data columns).
            output_path: Path for output file.

        Returns:
            str: Path to created file.
        """
        logger.info(f'Creating DATA file: {output_path}')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Data'

        # Get column order
        if self.column_order:
            codes = [code for code, desc in self.column_order]
            descs = [desc for code, desc in self.column_order]
        else:
            # Use DataFrame columns (excluding 'date')
            codes = [col for col in df.columns if col != 'date']
            descs = codes  # Use codes as descriptions if not available

        # Row 1: Column codes (empty first cell for date column)
        ws.cell(row=1, column=1, value='')
        for i, code in enumerate(codes, start=2):
            ws.cell(row=1, column=i, value=code)

        # Row 2: Column descriptions
        ws.cell(row=2, column=1, value='')
        for i, desc in enumerate(descs, start=2):
            ws.cell(row=2, column=i, value=desc)

        # Row 3+: Data rows
        for row_idx, (_, row) in enumerate(df.iterrows(), start=3):
            # Date in first column
            date_val = row['date']
            if isinstance(date_val, pd.Timestamp):
                date_str = date_val.strftime(config.DATE_FORMAT_OUTPUT)
            else:
                date_str = str(date_val)
            ws.cell(row=row_idx, column=1, value=date_str)

            # Data values
            for col_idx, code in enumerate(codes, start=2):
                if code in row:
                    value = row[code]
                    # Handle NA values
                    if pd.isna(value) or value == config.NA_VALUE:
                        ws.cell(row=row_idx, column=col_idx, value=config.NA_VALUE)
                    else:
                        ws.cell(row=row_idx, column=col_idx, value=value)
                else:
                    ws.cell(row=row_idx, column=col_idx, value=config.NA_VALUE)

        wb.save(output_path)
        wb.close()

        logger.info(f'DATA file created with {len(df)} data rows and {len(codes)} columns')

        return output_path

    def create_meta_file(self, output_path):
        """
        Create META Excel file with time series metadata.

        Args:
            output_path: Path for output file.

        Returns:
            str: Path to created file.
        """
        logger.info(f'Creating META file: {output_path}')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Metadata'

        # Header row
        headers = ['CODE', 'DESCRIPTION', 'FREQUENCY', 'UNIT_TYPE', 'DATA_TYPE', 'SOURCE', 'CATEGORY']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        # Data rows - one per time series
        if self.column_order:
            for row_idx, (code, desc) in enumerate(self.column_order, start=2):
                ws.cell(row=row_idx, column=1, value=code)
                ws.cell(row=row_idx, column=2, value=desc)
                ws.cell(row=row_idx, column=3, value=config.METADATA_DEFAULTS['FREQUENCY'])
                ws.cell(row=row_idx, column=4, value=config.METADATA_DEFAULTS['UNIT_TYPE'])
                ws.cell(row=row_idx, column=5, value=config.METADATA_DEFAULTS['DATA_TYPE'])
                ws.cell(row=row_idx, column=6, value=config.METADATA_DEFAULTS['SOURCE'])
                ws.cell(row=row_idx, column=7, value=config.METADATA_DEFAULTS['CATEGORY'])

        wb.save(output_path)
        wb.close()

        logger.info(f'META file created with {len(self.column_order) if self.column_order else 0} entries')

        return output_path

    def create_zip_file(self, data_file, meta_file, zip_path):
        """
        Create ZIP archive containing DATA and META files.

        Args:
            data_file: Path to DATA file.
            meta_file: Path to META file.
            zip_path: Path for ZIP file.

        Returns:
            str: Path to created ZIP file.
        """
        logger.info(f'Creating ZIP file: {zip_path}')

        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.write(data_file, os.path.basename(data_file))
            zf.write(meta_file, os.path.basename(meta_file))

        logger.info(f'ZIP file created')

        return zip_path

    def save_master_data(self, df):
        """
        Save updated master data file.

        Args:
            df: DataFrame with all data.

        Returns:
            str: Path to master file.
        """
        master_path = config.MASTER_DATA_FILE
        logger.info(f'Saving master data to: {master_path}')

        # Create master file with same structure as DATA file
        self.create_data_file(df, master_path)

        logger.info(f'Master data saved with {len(df)} rows')

        return master_path

    def generate_files(self, df):
        """
        Generate all output files.

        Creates:
        - Timestamped DATA, META, and ZIP files
        - Latest DATA, META, and ZIP files
        - Updated master data file

        Args:
            df: DataFrame with combined data.

        Returns:
            dict: Dictionary with paths to all created files.
        """
        logger.info('Generating output files')

        # Create directories
        ts_dir, latest_dir = self._create_output_dirs()

        # File names
        data_name = f'{config.DATA_FILE_PREFIX}_{self.timestamp}.xlsx'
        meta_name = f'{config.META_FILE_PREFIX}_{self.timestamp}.xlsx'
        zip_name = f'{config.ZIP_FILE_PREFIX}_{self.timestamp}.zip'

        latest_data_name = f'{config.DATA_FILE_PREFIX}_LATEST.xlsx'
        latest_meta_name = f'{config.META_FILE_PREFIX}_LATEST.xlsx'
        latest_zip_name = f'{config.ZIP_FILE_PREFIX}_LATEST.zip'

        # Timestamped paths
        ts_data_path = os.path.join(ts_dir, data_name)
        ts_meta_path = os.path.join(ts_dir, meta_name)
        ts_zip_path = os.path.join(ts_dir, zip_name)

        # Latest paths
        latest_data_path = os.path.join(latest_dir, latest_data_name)
        latest_meta_path = os.path.join(latest_dir, latest_meta_name)
        latest_zip_path = os.path.join(latest_dir, latest_zip_name)

        # Create timestamped files
        self.create_data_file(df, ts_data_path)
        self.create_meta_file(ts_meta_path)
        self.create_zip_file(ts_data_path, ts_meta_path, ts_zip_path)

        # Copy to latest
        shutil.copy2(ts_data_path, latest_data_path)
        shutil.copy2(ts_meta_path, latest_meta_path)
        shutil.copy2(ts_zip_path, latest_zip_path)
        logger.info('Copied files to latest folder')

        # Save master data
        master_path = self.save_master_data(df)

        result = {
            'data_file': ts_data_path,
            'meta_file': ts_meta_path,
            'zip_file': ts_zip_path,
            'latest_data': latest_data_path,
            'latest_meta': latest_meta_path,
            'latest_zip': latest_zip_path,
            'master_file': master_path,
            'timestamp': self.timestamp,
            'output_dir': ts_dir,
        }

        logger.info('All output files generated successfully')

        return result


def generate_files(df, column_order=None):
    """
    Convenience function to generate all output files.

    Args:
        df: DataFrame with combined data.
        column_order: Optional list of (code, description) tuples.

    Returns:
        dict: Paths to created files.
    """
    generator = BOACTARFileGenerator(column_order=column_order)
    return generator.generate_files(df)


if __name__ == '__main__':
    # Test the file generator
    from logger_setup import setup_logging
    from data_loader import load_data
    from parser import BOACTARParser

    setup_logging()

    # Load and parse data
    loaded_data = load_data()

    if loaded_data:
        parser = BOACTARParser()
        combined_df = parser.parse_and_merge(loaded_data)
        column_order = parser.get_column_order()

        # Generate files
        generator = BOACTARFileGenerator(column_order=column_order)
        result = generator.generate_files(combined_df)

        print('\nGenerated files:')
        for key, path in result.items():
            if isinstance(path, str) and os.path.exists(path):
                size = os.path.getsize(path)
                print(f'  {key}: {path} ({size:,} bytes)')
            else:
                print(f'  {key}: {path}')
    else:
        print('No data to process')
